from flask import Flask, render_template, url_for, redirect, send_file, request, flash
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
import requests
import numpy as np
from datetime import date
import zipfile
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import sqlite3

app = Flask(__name__)
app.secret_key = 'lxclnsdlwerjoi2394890875602j34lksdlrr902830jiajsdf908sd8f34jl35609spdj'

# hash password so it isn't stored in plain text
hash_pass = generate_password_hash('St@rTr3k$ucks!')

global last_execute
last_execute = 'November 17, 2021'


def get_date():
    # Find today's date to compare to emp lists
    today = date.today()
    today = today.strftime("%B %d, %Y")
    last_execute = today

    return last_execute


def overwrite_file(a, b):
    if a == b:
        pass
    else:
        temp = pd.read_excel('templates/Emp_List.xlsx')
        temp.to_excel('templates/Temp_Emp_List.xlsx', index=False)


def emp_by_client_id():
    headers = {
        'accept': 'application/json',
        'x-api-key': 'XreqEzU09o8GuvgCDR2yxagXBeYPEpzC8UBozYd6'
    }

    response = requests.get(
        'https://f9tornqbue.execute-api.us-west-2.amazonaws.com/prod/api/Employee/GetEmployeeListByClientId',
        headers=headers
    )

    response = response.json()
    response = pd.json_normalize(response)

    # Export list to excel file
    response.to_excel('templates/Emp_List.xlsx', index=False)
    set_col_width('templates/Emp_List.xlsx')


def compare_lists():
    # Read new and previous file
    df1 = pd.read_excel('templates/Emp_List.xlsx')
    df2 = pd.read_excel('templates/Temp_Emp_List.xlsx')

    # drop any NaN cells
    df1 = df1.fillna(0)
    df2 = df2.fillna(0)

    # try/except to catch API failure error
    try:
        # Put employeeID column in lists to compare
        df1_empIDs = df1['employeeID'].tolist()
        df2_empIDs = df2['employeeID'].tolist()
    except KeyError:
        print('KeyError: Possible API failure')
        raise

    # Iterate through lists looking for new additions
    new_empIDs = []
    for i in df1_empIDs:
        if i not in df2_empIDs:
            new_empIDs.append(i)
    while len(new_empIDs) == 0:
        print('No new hires found')
        new_empIDs.append('No new hires found')

    # Add new additions to list and export to excel
    df1_addition = []
    for i in new_empIDs:
        df1_addition.append(df1.loc[df1['employeeID'] == i])

    df1_addition = pd.DataFrame(np.concatenate(df1_addition))
    df1_addition = df1_addition.rename(
        columns={0: 'firstName', 1: 'lastName', 2: 'preferredName', 3: 'businessUnit',
                 4: 'jobTitle', 5: 'reportsTo', 6: 'employeeID', 7: 'username',
                 8: 'isActive', 9: 'peoHireDate', 10: 'erHireDate', 11: 'seniorityDate',
                 12: 'pobStatus', 13: 'pobStatusChangedDate', 14: 'lastDayWorked'})
    df1_addition.to_excel('templates/New_Hires.xlsx', index=False)
    set_col_width('templates/New_Hires.xlsx')



    # Create duplicate list to cut out any inactive employees
    df1_active = df1

    # Iterate through list, dropping any inactive employees
    temp = 0
    for i in df1_active['isActive']:
        if i == 'FALSE' or i == False:
            df1_active.drop(temp, inplace=True)
        temp += 1

    # Grab the only needed columns and export to excel
    contact_list = df1_active[['firstName', 'lastName', 'preferredName', 'businessUnit', 'jobTitle']]
    contact_list.to_excel('templates/contact_list.xlsx', index=False)
    set_col_width('templates/contact_list.xlsx')

    repairQ_file()


def repairQ_file():
    # define scope
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

    # add credentials to the account
    creds = ServiceAccountCredentials.from_json_keyfile_name('static/bold-maps-api-1562088768419-295fa4aed9b1.json',
                                                             scope)

    # authorize the clientsheet
    client = gspread.authorize(creds)

    # create instance of google sheet
    sheet = client.open_by_url(
        'https://docs.google.com/spreadsheets/d/1QMGiio57QR83DOn2PhB7NzmqahmjJ-UpgFSTSBAXJZw/edit?hl=en&forcehl=1#gid=0')

    # get first sheet
    sheet_instance = sheet.get_worksheet(0)
    store_nums = sheet_instance.col_values(1)
    store_names = sheet_instance.col_values(2)
    repairQ_codes = sheet_instance.col_values(4)
    store_ips = sheet_instance.col_values(7)

    # using a slice here to avoid grabbing column name on first row
    store_nums_data = store_nums[1:]
    store_names_data = store_names[1:]
    store_ips_data = store_ips[1:]

    # concatenate each list to form one column with store name and IP address
    store_data = [i + ':' + j for i, j in zip(store_names_data, store_ips_data)]

    # concatenate name to create username & email address
    new_hires = pd.read_excel('templates/New_Hires.xlsx')
    new_hires = new_hires.fillna(0)
    firstNames = new_hires['firstName'].tolist()
    lastNames = new_hires['lastName'].tolist()
    for i in firstNames:
        i.capitalize()
    for i in lastNames:
        i.capitalize()
    userNames = [i + '.' + j for i, j in zip(firstNames, lastNames)]

    ct = 0
    for i in userNames:
        for j in userNames[ct]:
            userNames[ct] = userNames[ct].replace("'", '')
            userNames[ct] = userNames[ct].replace(" ", '')
            userNames[ct] = userNames[ct].replace(" Jr", '')
        ct += 1

    domain = '@simplymac.com'
    emails = []
    for i in userNames:
        emails.append(i + domain)

    job_title = new_hires['jobTitle'].tolist()

    # extract store number from new_hires sheet in order to put relevant info for each new hire
    nums = new_hires['businessUnit'].tolist()
    nums_extracted = []
    for i in nums:
        nums_extracted.append(i[:3])

    # append needed IPs to temporary list
    temp_IPs = []
    location_strings = [str(i) for i in nums_extracted]

    for i in location_strings:
        for k in store_data:
            if i in k[:3]:
                temp_IPs.append(k)

    # add necessary RepairQ
    loc_advisors = 'UT101 - Orem:regional_manager|UT102 - Fort Union:regional_manager|UT105 - St. George:regional_manager|' \
                   'UT106 - Foothill:regional_manager|ID401 - Idaho Falls:regional_manager|' \
                   'MO510 - Springfield MO:regional_manager|TX203 - Tyler:regional_manager|OR410 - Bend:regional_manager|' \
                   'TX200 - Lubbock:regional_manager|MT404 - Missoula:regional_manager|MT406 - Bozeman:regional_manager|' \
                   'UT100 - Park City:regional_manager|AR205 - Rogers:regional_manager|TN312 - Nashville:regional_manager|' \
                   'TN313 - Vanderbilt:regional_manager|TN314 - Chattanooga:regional_manager|TN315 - Johnson City:regional_manager|' \
                   'KY316 - Louisville:regional_manager|IN305 - Ft. Wayne:regional_manager|IN301 - Evansville:regional_manager|' \
                   'GA702 - Athens:regional_manager|VA704 - Charlottesville:regional_manager|GA705 - Macon:regional_manager|' \
                   'GA706 - West Cobb:regional_manager|GA707 - Peachtree City:regional_manager|' \
                   'GA708 - Norcross:regional_manager|GA710 - Forsyth:regional_manager|' \
                   'GA711 - East Cobb:regional_manager|AL712 - Montgomery:regional_manager|' \
                   'GA713 - St. Simons:regional_manager|' \
                   'OR419 - Corvallis:regional_manager|OR420 - Eugene:regional_manager|OH308 - Cincinnati:regional_manager|' \
                   'TX206 - Katy:regional_manager|FL917 - Miami Warehouse:regional_manager|FL714 - Orlando:regional_manager|' \
                   'FL715 - Sawgrass:regional_manager|FL716 - Dolphin:regional_manager|NC317 - Asheville:regional_manager|' \
                   'GA717 - Gainesville:regional_manager|OR421 - Hillsboro:regional_manager|FL801 - Destin:regional_manager|' \
                   'TX207 - Waco:regional_manager|MO514 - Columbia:regional_manager|FL805 - Orlando UCF:regional_manager|' \
                   'TX208 - San Marcos:regional_manager|FL802 - Tallahassee:regional_manager|NC605 - Winston Salem:regional_manager|' \
                   'SC604 - Myrtle Beach:regional_manager|AL803 - Tuscaloosa:regional_manager|KS511 - Lawrence:regional_manager|' \
                   'SC602 - Columbia SC:regional_manager|NC603 - Wilmington:regional_manager|SC601 - Mt. Pleasant:regional_manager|' \
                   'KS513 - Wichita:regional_manager|FL807 - Clearwater:regional_manager|FL804 - Jacksonville:regional_manager'

    specialist = ':sales_manager'
    tech = ':technical_manager,sales_manager'
    sr_specialist = ':regional_manager,shift_lead,technical_manager,sales_manager'
    sr_tech = ':shift_lead,technical_manager,sales_manager,regional_manager'
    ops_mgr = ':inventory_manager,shift_lead,technical_manager,sales_manager,regional_manager'
    store_ldr = ':inventory_manager,shift_lead,technical_manager,sales_manager,regional_manager'
    mkt_director = ':inventory_manager,shift_lead,technical_manager,sales_manager,regional_manager'

    state_codes = loc_advisors
    loc_roles = []

    counter = 0
    for i in location_strings:
        for j in repairQ_codes:
            if i in j[:5]:
                if job_title[counter] == 'Apple Specialist':
                    full_code = j + specialist
                    loc_roles.append(full_code)
                elif job_title[counter] == 'Senior Apple Specialist':
                    full_code = j + sr_specialist
                    loc_roles.append(full_code)
                elif job_title[counter] == 'Technician':
                    full_code = j + tech
                    loc_roles.append(full_code)
                else:
                    temp = str(j + ':regional_manager')
                    full_code = state_codes.replace(temp, j + '{position}')
                    loc_roles.append(full_code)

                counter += 1

    counter2 = 0
    for i in job_title:
        if i == 'Service Leader' or i == 'Service Manager':
            loc_roles[counter2] = loc_roles[counter2].format(position=sr_tech)
        elif i == 'Operation Manager' or i == 'Assistant Manager' or i == 'ASM':
            loc_roles[counter2] = loc_roles[counter2].format(position=ops_mgr)
        elif i == 'Store Leader' or i == 'Store Manager':
            loc_roles[counter2] = loc_roles[counter2].format(position=store_ldr)

        counter2 += 1

    # set cell values for remaining columns that do not change from employee to employee
    active_col = []
    role_type_col = []
    contact_meth_col = []
    timesheet_col = []
    discount_col = []
    discount_col2 = []

    # input static values based on length of other columns
    for i in range(len(temp_IPs)):
        active_col.append('yes')
        role_type_col.append('User')
        contact_meth_col.append('Email')
        timesheet_col.append('no')
        discount_col.append('None')
        discount_col2.append('None')

    # building dataframe
    store_data_df = pd.DataFrame({'First Name': firstNames, 'Last Name': lastNames, 'Username': userNames,
                                  'Email': emails, 'Location': nums_extracted, 'Role': job_title, 'Ips': temp_IPs,
                                  'Location Roles': loc_roles, 'Active': active_col, 'Role Type': role_type_col,
                                  'Contact Method': contact_meth_col, 'Timesheet Required': timesheet_col,
                                  'Discount Limit': discount_col, 'Discount Limit Type': discount_col2})

    store_data_df.to_excel('templates/repairQ_import.xlsx', index=False)

    zip_files()


def set_col_width(x):
    # Load excel sheet created in first method
    wb = load_workbook(x)
    ws = wb['Sheet1']

    # Find row with most characters and set column width
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

    ws.column_dimensions = dim_holder

    wb.save(filename=x)


def zip_files():
    # Because Brian and I are lazy, set column width to match longest cell value
    set_col_width('templates/Emp_List.xlsx')
    set_col_width('templates/New_Hires.xlsx')
    set_col_width('templates/contact_list.xlsx')
    set_col_width('templates/repairQ_import.xlsx')

    # Create zip file file with all excel files
    files_to_zip = ['templates/Emp_List.xlsx', 'templates/New_Hires.xlsx', 'templates/contact_list.xlsx',
                    'templates/repairQ_import.xlsx']
    with zipfile.ZipFile('templates/Emp_List&New_Hires.zip', 'w') as zipF:
        for file in files_to_zip:
            zipF.write(file, compress_type=zipfile.ZIP_DEFLATED)


@app.route('/home', methods=['GET', 'POST'])
def home():
    try:
        return render_template('home.html', last_execute=last_execute)
    except Exception as e:
        return str(e)


@app.route('/')
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form['username'] != 'SimplyMacIT' or request.form['password'] != hash_pass:
            error = 'Invalid login credentials. Please try again.'
        else:
            return render_template('home.html')
    return render_template('login.html', error=error)


@app.route('/background_emp_list')
def background_emp_list():
    emp_by_client_id()
    compare_lists()
    try:
        return 'nothing'
    except Exception as e:
        return str(e)


@app.route('/background_overwrite')
def background_overwrite():
    df1 = pd.read_excel('templates/Emp_List.xlsx')
    df2 = pd.read_excel('templates/Temp_Emp_List.xlsx')
    df1_empIDs = df1['employeeID'].tolist()
    df2_empIDs = df2['employeeID'].tolist()
    overwrite_file(df1_empIDs, df2_empIDs)
    try:
        return 'nothing'
    except Exception as e:
        return str(e)


@app.route('/download_file')
def download_file():
    path = 'templates/Emp_List&New_Hires.zip'
    try:
        return send_file(path, as_attachment=True)
    except Exception as e:
        return str(e)


if __name__ == '__main__':
    app.run(debug=True, host='192.168.10.100')
